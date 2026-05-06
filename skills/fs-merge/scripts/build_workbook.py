#!/usr/bin/env python3
"""
Build merged FS workbook in the example layout (v2).

Layout (single sheet "FS"):
  Col A: section markers (' - ')
  Col B: section names (PL/MS/BS/CHECK/Ref) and 구분 / 계정과목
  Col C: 실계정 (0 = subtotal, 1 = leaf, blank = section header)
  Col D: Group (BS only; left blank for manual fill)
  Col E~ : year columns (date headers)
  Then flag columns (재계산_차이 / OCR_의심 / 매칭_의심)

Optional period_notes row appears immediately below the date header row
when any non-12-month period exists.

Usage:
  python build_workbook.py <input.json> <output.xlsx>

Input JSON schema (v2):
{
  "fiscal_year_ends": ["2022-12-31", ...],   # ISO dates, oldest -> newest
  "period_notes": [null, null, null, "6M interim"],   # optional, parallel
  "statements": {
    "PL": [
      {
        "account": "Ⅰ.매출액",
        "is_leaf": false,
        "is_section_header": false,            # optional, default false
        "values": [null, 124969182895, ...],   # one per fiscal_year_ends
        "flags": {                              # optional
          "재계산_차이": ["2023: 12,345,000 vs 12,344,000 (1,000)", null, ...],
          "OCR_의심": [null, ...],
          "매칭_의심": [null, ...]
        }
      },
      ...
    ],
    "MS": [...],
    "BS": [   # BS rows: same as above, plus optional "group" field
      {"account": "...", "is_leaf": false, "group": "", "values": [...], "flags": {...}},
      ...
    ]
  },
  "checks": {              # optional, two formats supported per check
    "PL": [
      # Format 1 — legacy positive sum:
      {"label": "Ⅳ.판매관리비 = Σ leaves",
       "subtotal_idx": 8,
       "child_indices": [9, 10, 11, ...]},

      # Format 2 — signed terms (linear combination, for non-additive subtotals):
      {"label": "Ⅲ.매출총이익 = Ⅰ - Ⅱ",
       "subtotal_idx": 7,
       "terms": [{"idx": 0, "sign": 1},
                 {"idx": 2, "sign": -1}]},
      ...
    ],
    "MS": [...],
    "BS": [...]
  },
  "refs": [                # optional cross-statement references
    {"label": "당기제품제조원가", "stmt_a": "PL", "row_idx_a": 5,
                                  "stmt_b": "MS", "row_idx_b": 41}
  ]
}

The build script tracks where each statement-row lands in the workbook so
formulas reference the right cells.

v2 changes (vs v1):
- checks[].terms: signed linear combination support for non-additive subtotals
  (매출총이익, 영업이익, 소득세차감전이익, 당기순이익, 상품매출원가 etc.)
- is_section_header: leaves the 실계정 column blank for visual divider rows
  like [자산]/[부채]/[자본]
- period_notes: optional sub-header row for non-12M periods
- freeze_panes computed dynamically based on first PL data row
"""
from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# Output layout constants
SECTION_ORDER = ["PL", "MS", "BS"]
DATA_START_COL = 5  # column E
FLAG_NAMES = ["재계산_차이", "OCR_의심", "매칭_의심"]


def _iso_to_date(s: str) -> datetime:
    return datetime.strptime(s, "%Y-%m-%d")


def _write_section_header(ws, row: int, name: str, n_years: int, has_group: bool):
    """Write the marker row and the column header row. Returns the header row index."""
    ws.cell(row=row, column=1, value=" - ")
    ws.cell(row=row, column=2, value=name)
    ws.cell(row=row, column=2).font = Font(bold=True)

    header_row = row + 2  # blank row between marker and header
    ws.cell(row=header_row, column=2, value="구분").font = Font(bold=True)
    ws.cell(row=header_row, column=3, value="실계정").font = Font(bold=True)
    if has_group:
        ws.cell(row=header_row, column=4, value="Group").font = Font(bold=True)
    return header_row


def _write_year_headers(ws, header_row: int, fiscal_year_ends: list[str]):
    """Write year-end dates as column headers starting at DATA_START_COL (E)."""
    for i, iso in enumerate(fiscal_year_ends):
        col = DATA_START_COL + i
        c = ws.cell(row=header_row, column=col, value=_iso_to_date(iso))
        c.number_format = "yyyy-mm-dd"
        c.font = Font(bold=True)


def _write_period_notes(ws, note_row: int, period_notes: list, n_years: int):
    """Write optional period note labels (e.g., '6M interim') beneath date headers."""
    for i, note in enumerate(period_notes):
        if note and i < n_years:
            col = DATA_START_COL + i
            c = ws.cell(row=note_row, column=col, value=str(note))
            c.font = Font(italic=True, size=9)
            c.alignment = Alignment(horizontal="center")


def _write_flag_headers(ws, header_row: int, n_years: int):
    """Write flag column headers after the year columns."""
    base = DATA_START_COL + n_years
    for i, name in enumerate(FLAG_NAMES):
        c = ws.cell(row=header_row, column=base + i, value=name)
        c.font = Font(bold=True)


def _excel_col(idx: int) -> str:
    """1-based column index -> letter."""
    return get_column_letter(idx)


def _write_data_rows(ws, start_row: int, rows: list[dict], n_years: int,
                     has_group: bool) -> tuple[int, list[int]]:
    """Write data rows. Returns (next_row, list_of_excel_row_per_input_row).

    Year columns always start at column E (DATA_START_COL). The Group
    column at D is reserved: filled for BS rows, left blank for PL/MS.

    Section header rows (is_section_header: true) leave column C blank.
    """
    flag_base = DATA_START_COL + n_years

    excel_rows: list[int] = []
    for i, r in enumerate(rows):
        row = start_row + i
        excel_rows.append(row)

        ws.cell(row=row, column=2, value=r["account"])

        if r.get("is_section_header"):
            # Leave 실계정 column (C) blank for visual divider rows
            pass
        else:
            ws.cell(row=row, column=3, value=0 if not r.get("is_leaf") else 1)

        if has_group:
            ws.cell(row=row, column=4, value=r.get("group", ""))

        # values
        for j, v in enumerate(r.get("values", [])):
            if j >= n_years:
                break
            col = DATA_START_COL + j
            if v is None or v == "":
                continue
            c = ws.cell(row=row, column=col, value=v)
            c.number_format = "#,##0;(#,##0);-"

        # flags
        flags = r.get("flags", {}) or {}
        for k, name in enumerate(FLAG_NAMES):
            arr = flags.get(name, [])
            notes = []
            for j, note in enumerate(arr or []):
                if note:
                    notes.append(str(note))
            if notes:
                ws.cell(row=row, column=flag_base + k, value="; ".join(notes))

    return start_row + len(rows), excel_rows


def _resolve_terms(chk: dict) -> list[dict] | None:
    """Resolve a check spec into a list of {idx, sign} terms.

    Supports two formats:
      1. {"child_indices": [a, b, c]}  → all sign=+1 (legacy)
      2. {"terms": [{"idx": a, "sign": 1}, ...]}  → arbitrary signs

    Returns None if neither format is provided or the spec is empty.
    """
    if "terms" in chk and chk["terms"]:
        return [
            {"idx": int(t["idx"]), "sign": 1 if int(t.get("sign", 1)) >= 0 else -1}
            for t in chk["terms"]
        ]
    if "child_indices" in chk and chk["child_indices"]:
        return [{"idx": int(c), "sign": 1} for c in chk["child_indices"]]
    return None


def _build_check_formula(letter: str, sub_row: int,
                         terms: list[dict], excel_rows: list[int]) -> str:
    """Build the Excel formula string for one check cell.

    All-positive contiguous → SUM range form.
    Otherwise → explicit signed expression.
    """
    if not terms:
        return f"={letter}{sub_row}"

    all_pos = all(t["sign"] >= 0 for t in terms)
    term_rows = [excel_rows[t["idx"]] for t in terms]
    contiguous = (
        len(term_rows) > 1
        and term_rows == list(range(min(term_rows), max(term_rows) + 1))
    )

    if all_pos and contiguous:
        rng = f"{letter}{min(term_rows)}:{letter}{max(term_rows)}"
        return f"={letter}{sub_row}-SUM({rng})"

    # Build signed expression term by term
    parts: list[str] = []
    for t in terms:
        cell = f"{letter}{excel_rows[t['idx']]}"
        if not parts:
            parts.append(cell if t["sign"] >= 0 else f"-{cell}")
        else:
            parts.append(f"+{cell}" if t["sign"] >= 0 else f"-{cell}")
    rhs = "".join(parts)
    return f"={letter}{sub_row}-({rhs})"


def _write_check_section(ws, start_row: int, stmt_name: str,
                         checks: list[dict], excel_rows: list[int],
                         n_years: int) -> int:
    """Write CHECK rows. Each check evaluates `subtotal - linear_combo(terms)`.

    For correctly-balanced statements, every cell should display 0 (rendered
    as '-' due to the number format).
    """
    if not checks:
        return start_row

    ws.cell(row=start_row, column=2, value="CHECK").font = Font(bold=True, italic=True)

    next_row = start_row + 1
    for chk in checks:
        ws.cell(row=next_row, column=2, value=chk.get("label", ""))
        sub_idx = chk.get("subtotal_idx")
        if sub_idx is None or sub_idx >= len(excel_rows):
            next_row += 1
            continue
        sub_row = excel_rows[sub_idx]

        terms = _resolve_terms(chk)
        if terms is None:
            terms = []

        # Validate term indices
        if any(t["idx"] >= len(excel_rows) for t in terms):
            next_row += 1
            continue

        for j in range(n_years):
            col = DATA_START_COL + j
            letter = _excel_col(col)
            formula = _build_check_formula(letter, sub_row, terms, excel_rows)
            c = ws.cell(row=next_row, column=col, value=formula)
            c.number_format = "#,##0;(#,##0);-"
        next_row += 1
    return next_row + 1


def _write_ref_section(ws, start_row: int, refs: list[dict],
                       all_excel_rows: dict[str, list[int]],
                       n_years: int) -> int:
    """Cross-statement reference rows."""
    if not refs:
        return start_row
    ws.cell(row=start_row, column=2, value="Ref").font = Font(bold=True, italic=True)
    next_row = start_row + 1
    for ref in refs:
        a, b = ref["stmt_a"], ref["stmt_b"]
        if a not in all_excel_rows or b not in all_excel_rows:
            next_row += 1
            continue
        ra = all_excel_rows[a][ref["row_idx_a"]]
        rb = all_excel_rows[b][ref["row_idx_b"]]
        ws.cell(row=next_row, column=2, value=ref.get("label", ""))
        for j in range(n_years):
            col = DATA_START_COL + j
            letter = _excel_col(col)
            formula = f"={letter}{ra}-{letter}{rb}"
            c = ws.cell(row=next_row, column=col, value=formula)
            c.number_format = "#,##0;(#,##0);-"
        next_row += 1
    return next_row + 1


def build(payload: dict, output_path: Path) -> None:
    fye = payload["fiscal_year_ends"]
    period_notes = payload.get("period_notes") or []
    n_years = len(fye)
    statements = payload["statements"]
    checks = payload.get("checks", {}) or {}
    refs = payload.get("refs", []) or []

    has_period_notes = bool(period_notes) and any(period_notes)

    wb = Workbook()
    ws = wb.active
    ws.title = "FS"

    # Column widths matching example
    ws.column_dimensions["A"].width = 2.7
    ws.column_dimensions["B"].width = 27
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    for i in range(n_years + len(FLAG_NAMES)):
        ws.column_dimensions[_excel_col(DATA_START_COL + i)].width = 18

    # B1 marker
    ws.cell(row=1, column=2, value="FS").font = Font(bold=True)

    all_excel_rows: dict[str, list[int]] = {}
    cur_row = 4
    first_data_row: int | None = None

    for stmt in SECTION_ORDER:
        rows = statements.get(stmt, [])
        if not rows:
            continue
        has_group = (stmt == "BS")

        # Section marker + header row
        header_row = _write_section_header(ws, cur_row, stmt, n_years, has_group)
        _write_year_headers(ws, header_row, fye)
        _write_flag_headers(ws, header_row, n_years)

        # Optional period_notes row beneath date header
        if has_period_notes:
            note_row = header_row + 1
            _write_period_notes(ws, note_row, period_notes, n_years)
            data_start = note_row + 1
        else:
            data_start = header_row + 1

        if first_data_row is None:
            first_data_row = data_start

        # Data rows
        next_row, excel_rows = _write_data_rows(
            ws, data_start, rows, n_years, has_group
        )
        all_excel_rows[stmt] = excel_rows

        # CHECK section (blank line, then label, then check rows)
        next_row = _write_check_section(
            ws, next_row + 1, stmt, checks.get(stmt, []),
            excel_rows, n_years
        )

        cur_row = next_row + 2

    # Refs at the very end
    if refs:
        cur_row = _write_ref_section(
            ws, cur_row, refs, all_excel_rows, n_years
        )

    # Freeze panes at the first data row of the first statement
    if first_data_row is not None:
        ws.freeze_panes = f"E{first_data_row}"

    wb.save(output_path)


def main():
    if len(sys.argv) != 3:
        print("Usage: build_workbook.py <input.json> <output.xlsx>", file=sys.stderr)
        sys.exit(2)
    with open(sys.argv[1], "r", encoding="utf-8") as f:
        payload = json.load(f)
    out = Path(sys.argv[2])
    out.parent.mkdir(parents=True, exist_ok=True)
    build(payload, out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
